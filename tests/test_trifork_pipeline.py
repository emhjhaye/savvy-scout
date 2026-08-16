from datetime import datetime, timezone

from openpyxl import load_workbook

from savvy_scout.escalation.brief import record_brief
from savvy_scout.export.trifork_pipeline import HEADERS, TOTAL_COLUMNS, update_trifork_pipeline


def _notice(conn, ref, target, actor="Mark", rating="FLAG"):
    now = datetime.now(timezone.utc).isoformat()
    status = "ESCALATED_TO_VICTORIA" if target == "ESCALATED_TO_VICTORIA" else "REJECTED"
    notice_id = conn.execute(
        "INSERT INTO notices (ref, title, buyer, source, uk_stage, status, sector, owner, "
        "indicative_value, cpv_primary, deadline, text_blob, first_seen_at, last_swept_at, "
        "created_at, updated_at, raw_json) VALUES (?, ?, 'Buyer', 'Find a Tender', 'UK3', ?, "
        "'Fintech', 'Mark', 'GBP 100000', '72200000', '2026-08-20', 'Requirement', ?, ?, ?, ?, '{}')",
        (ref, f"Title {ref}", status, now, now, now, now),
    ).lastrowid
    conn.execute(
        "INSERT INTO status_history (notice_id, from_status, to_status, changed_by, changed_at, reason) "
        "VALUES (?, 'AWAITING_PHASE2_APPROVAL', ?, ?, ?, 'Owner decision')",
        (notice_id, target, actor, now),
    )
    if rating:
        conn.execute(
            "INSERT INTO phase2_assessments (notice_id, capability_fit_rating, capability_fit_reasoning, "
            "competitor_position_rating, competitor_position_reasoning, right_to_win_rating, "
            "right_to_win_reasoning, overall_rating, overall_reasoning, open_questions, model_used, created_at) "
            "VALUES (?, 'HIGH', 'Strong fit', 'UNKNOWN', 'Unknown', 'MED', 'Plausible', ?, "
            "'Owner-reviewed recommendation', '[]', 'test', ?)",
            (notice_id, rating, now),
        )
    conn.commit()
    return notice_id


def test_tracker_preserves_template_and_upserts_by_notice_reference(conn, tmp_path):
    _notice(conn, "REF-PASS", "ESCALATED_TO_VICTORIA", rating="PURSUE")
    _notice(conn, "REF-FLAG", "ESCALATED_TO_VICTORIA", rating="FLAG")
    _notice(conn, "REF-FAIL", "REJECTED", rating="DECLINE")
    _notice(conn, "REF-SYSTEM", "REJECTED", actor="system_cleanup")
    _notice(conn, "REF-NO-ASSESSMENT", "REJECTED", rating=None)
    output = tmp_path / "My_Trifork_Pipeline_Tracker.xlsx"

    first = update_trifork_pipeline(conn, str(output))
    second = update_trifork_pipeline(conn, str(output))

    assert first["inserted"] == 3
    assert second["inserted"] == 0
    assert second["updated"] == 3
    workbook = load_workbook(output, data_only=False)
    assert tuple(workbook["Flag"].cell(2, column).value for column in range(1, TOTAL_COLUMNS + 1)) == HEADERS
    assert workbook["Master Summary"]["B4"].value == "=COUNTA(Pass!A3:A1000)"
    references = {
        sheet_name: [workbook[sheet_name].cell(row, 1).value for row in range(3, workbook[sheet_name].max_row + 1)
                     if workbook[sheet_name].cell(row, 1).value]
        for sheet_name in ("Pass", "Flag", "Fail")
    }
    assert references["Pass"] == ["REF-PASS"]
    assert references["Flag"] == ["REF-FLAG"]
    assert references["Fail"] == ["REF-FAIL"]
    assert "REF-SYSTEM" not in references["Pass"] + references["Flag"] + references["Fail"]
    assert "REF-NO-ASSESSMENT" not in references["Pass"] + references["Flag"] + references["Fail"]


def test_tracker_links_to_recorded_addendum_and_brief(conn, tmp_path):
    notice_id = _notice(conn, "REF-LINKED", "ESCALATED_TO_VICTORIA", rating="FLAG")
    tracker_dir = tmp_path / "Pipeline Tracker"
    artifacts_dir = tmp_path / "Addendum and Brief per Phase 2 Pass & Flag Opportunities" / "01. Some Title"
    artifacts_dir.mkdir(parents=True)
    addendum_path = artifacts_dir / "Some_Title_Internal_Addendum.docx"
    addendum_path.write_bytes(b"fake docx")
    record_brief(conn, notice_id, "escalated", str(addendum_path), "test", "INTERNAL_ADDENDUM")
    output = tracker_dir / "My_Trifork_Pipeline_Tracker.xlsx"

    update_trifork_pipeline(conn, str(output))

    workbook = load_workbook(output)
    sheet = workbook["Flag"]
    addendum_cell = sheet.cell(3, 20)
    assert addendum_cell.value == "Open Internal Addendum"
    assert addendum_cell.hyperlink is not None
    assert "Internal_Addendum.docx" in addendum_cell.hyperlink.target
    # Regression (2026-08-16): the hyperlink was set but never visually
    # styled as a link (still plain black, no underline), so it looked like
    # inert text in Excel even though it was clickable.
    assert addendum_cell.font.color.rgb in ("000563C1", "FF0563C1", "0563C1")
    assert addendum_cell.font.underline == "single"
    capture_cell = sheet.cell(3, 21)
    assert capture_cell.value == "Not yet generated"


def test_zero_value_is_presented_as_not_stated(conn, tmp_path):
    notice_id = _notice(conn, "REF-ZERO", "REJECTED", rating="DECLINE")
    conn.execute("UPDATE notices SET indicative_value = '0.0 GBP' WHERE id = ?", (notice_id,))
    conn.commit()
    output = tmp_path / "tracker.xlsx"

    update_trifork_pipeline(conn, str(output))

    workbook = load_workbook(output)
    assert workbook["Fail"]["I3"].value == "Not stated"
