from openpyxl import load_workbook

from savvy_scout.export.excel_tracker import COLUMNS, export_tracker
from savvy_scout.models.notice import Notice
from savvy_scout.sources.ocds_parser import ParsedNotice
from savvy_scout.sweep.dedupe import upsert_notice
from savvy_scout.triage.gates import triage_notice


def _make_parsed(ref: str, title: str, buyer: str, text_blob: str, cpv_primary: str | None = None) -> ParsedNotice:
    notice = Notice(
        ref=ref,
        title=title,
        buyer=buyer,
        source="Find a Tender",
        notice_type="UK3",
        uk_stage="UK3",
        raw_json="{}",
        cpv_primary=cpv_primary,
    )
    return ParsedNotice(notice=notice, text_blob=text_blob, tender_status="active")


def test_export_creates_expected_sheets_and_columns(conn, tmp_path):
    # A clean pass on every gate: Fintech/Mark, bespoke build, direct award
    # language (Gate 3), a Gate 5 PASS-list CPV code, no value (Filter 3 n/a).
    pass_notice = _make_parsed(
        "REF-PASS", "Real Time Payments Platform", "Some Bank",
        "bespoke build of a real-time payments platform for a bank, a direct award open tender",
        cpv_primary="72200000",
    )
    pass_id = upsert_notice(conn, pass_notice)
    triage_notice(conn, pass_id)

    # Sector matches cleanly (Energy/Mark) so Gate 1 passes, but Gate 2 fails
    # on hardware/resale language, which becomes the headline.
    fail_notice = _make_parsed(
        "REF-FAIL", "Server Hardware Resale", "National Grid",
        "hardware resale of legacy servers for the energy network",
    )
    fail_id = upsert_notice(conn, fail_notice)
    triage_notice(conn, fail_id)

    output_path = tmp_path / "tracker.xlsx"
    export_tracker(conn, str(output_path))

    wb = load_workbook(output_path)
    assert "Phase 1 - Flags" in wb.sheetnames
    assert "To review" in wb.sheetnames
    assert "Closed or awarded" in wb.sheetnames
    assert "Out of scope - no owner" in wb.sheetnames
    assert "Phase 2 - Pipeline" in wb.sheetnames
    assert "Legend and method" in wb.sheetnames

    for sheet_name in ["Phase 1 - Flags", "To review", "Closed or awarded", "Out of scope - no owner"]:
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        assert header == COLUMNS

    to_review_ws = wb["To review"]
    refs_to_review = [row[0].value for row in to_review_ws.iter_rows(min_row=2)]
    assert "REF-PASS" in refs_to_review

    out_of_scope_ws = wb["Out of scope - no owner"]
    refs_out_of_scope = [row[0].value for row in out_of_scope_ws.iter_rows(min_row=2)]
    assert "REF-FAIL" in refs_out_of_scope
