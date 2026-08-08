"""Regression test (SPEC.md A5): compares the engine's Phase 1 outcomes
against a provided baseline tracker Excel file, row by row, for a given
notice-ref set, listing every disagreement.

The baseline file (the manual 129-notice sweep) isn't in the project folder
yet, so the reader below is deliberately tolerant of column naming: it looks
for a column containing "ref" for the join key and a column containing
"triage status" (or "outcome" / "verdict") for the human-assigned outcome,
rather than hardcoding exact header text. Once the real file is supplied,
tighten this if the headers turn out to need it."""

import sqlite3

from openpyxl import Workbook, load_workbook


def _find_column(header_row: list, *keywords: str) -> int | None:
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if any(keyword in text for keyword in keywords):
            return idx
    return None


def read_baseline(path: str, sheet_name: str | None = None) -> dict[str, str]:
    """Returns {ref: human_outcome_raw_text}."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {}

    ref_col = _find_column(list(header), "ref")
    outcome_col = _find_column(list(header), "triage status", "outcome", "verdict")
    if ref_col is None or outcome_col is None:
        raise ValueError(
            "Could not find a ref column and a triage status/outcome column in "
            f"{path}, sheet {ws.title}. Found headers: {header}"
        )

    baseline = {}
    for row in rows:
        ref = row[ref_col]
        outcome = row[outcome_col]
        if ref is None:
            continue
        baseline[str(ref).strip()] = str(outcome).strip() if outcome is not None else ""
    return baseline


def normalise_outcome(raw: str) -> str:
    text = (raw or "").strip().lower()
    if "pass to" in text:
        return "FLAG"
    if "fail" in text:
        return "FAIL"
    if "flag" in text:
        return "FLAG"
    if "monitor" in text:
        return "MONITOR"
    if "pass" in text:
        return "PASS"
    return "OTHER"


def compare_against_baseline(conn: sqlite3.Connection, baseline: dict[str, str]) -> list[dict]:
    """Returns one row per baseline ref: machine outcome (from the latest
    triage run for that ref, if any), human outcome, and whether they agree."""
    diff_rows = []
    for ref, human_raw in baseline.items():
        notice = conn.execute("SELECT id, title FROM notices WHERE ref = ?", (ref,)).fetchone()
        if notice is None:
            diff_rows.append(
                {
                    "ref": ref,
                    "title": "",
                    "machine_outcome": "NOT IN DATABASE",
                    "human_outcome_raw": human_raw,
                    "human_outcome": normalise_outcome(human_raw),
                    "agree": False,
                    "notes": "Notice not swept/found in the database for this date range.",
                }
            )
            continue

        run = conn.execute(
            "SELECT headline_gate, headline_outcome, headline_reason FROM triage_runs "
            "WHERE notice_id = ? ORDER BY id DESC LIMIT 1",
            (notice["id"],),
        ).fetchone()

        if run is None:
            diff_rows.append(
                {
                    "ref": ref,
                    "title": notice["title"],
                    "machine_outcome": "NOT TRIAGED",
                    "human_outcome_raw": human_raw,
                    "human_outcome": normalise_outcome(human_raw),
                    "agree": False,
                    "notes": "Notice is in the database but has no triage run.",
                }
            )
            continue

        machine_outcome = run["headline_outcome"]
        human_outcome = normalise_outcome(human_raw)
        diff_rows.append(
            {
                "ref": ref,
                "title": notice["title"],
                "machine_outcome": machine_outcome,
                "human_outcome_raw": human_raw,
                "human_outcome": human_outcome,
                "agree": machine_outcome == human_outcome,
                "notes": run["headline_reason"] if machine_outcome != human_outcome else "",
            }
        )
    return diff_rows


def write_diff_report(diff_rows: list[dict], output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Regression diff"
    ws.append(
        ["Ref", "Title", "Machine outcome", "Human outcome (raw)", "Human outcome (normalised)", "Agree", "Notes"]
    )
    for row in diff_rows:
        ws.append(
            [
                row["ref"],
                row["title"],
                row["machine_outcome"],
                row["human_outcome_raw"],
                row["human_outcome"],
                "Yes" if row["agree"] else "No",
                row["notes"],
            ]
        )

    disagreements = [r for r in diff_rows if not r["agree"]]
    summary = wb.create_sheet(title="Summary")
    summary.append(["Total baseline rows", len(diff_rows)])
    summary.append(["Agreements", len(diff_rows) - len(disagreements)])
    summary.append(["Disagreements", len(disagreements)])

    wb.save(output_path)
    return output_path


def run_regression_test(conn: sqlite3.Connection, baseline_path: str, output_path: str) -> dict:
    baseline = read_baseline(baseline_path)
    diff_rows = compare_against_baseline(conn, baseline)
    write_diff_report(diff_rows, output_path)
    disagreements = [r for r in diff_rows if not r["agree"]]
    return {
        "total": len(diff_rows),
        "agreements": len(diff_rows) - len(disagreements),
        "disagreements": len(disagreements),
        "output_path": output_path,
    }
