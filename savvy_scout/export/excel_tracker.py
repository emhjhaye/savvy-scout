"""Excel tracker export (SPEC.md A4). Sheets and columns match SPEC.md
literally. SPEC.md doesn't define exactly which rows land on "Phase 1 -
Flags" vs "To review" vs "Handoffs", so this is Phase A's working split,
documented here and on the "Legend and method" sheet so it's easy to correct:

- Phase 1 - Flags: headline outcome FLAG or MAYBE (a judgement call is needed).
- To review: headline outcome PASS or INFERRED_FIT, in the Phase 2 pipeline
  (queued for or awaiting the owner's Phase 2 approval click). PASS/FLAG/
  MAYBE all skip the Phase 1 owner queue and go straight to the automated
  Phase 2 scope read, so this is where an owner actually finds them now.
- Handoffs - <owner>: one sheet per owner, the union of that owner's rows
  from Flags and To review, i.e. everything currently on their desk.
- Closed or awarded: headline outcome FAIL from Gate 4 (window) specifically.
- Out of scope - no owner: headline outcome FAIL from any other gate.
- Phase 2 - Pipeline: any notice from PHASE2_SCOPED onward, regardless of
  headline outcome.
- Legend and method: static reference sheet.
"""

import json
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

COLUMNS = [
    "Ref",
    "Date spotted",
    "Opportunity title",
    "Buyer",
    "Owner",
    "Source",
    "Indicative value",
    "CPV codes",
    "UK stage",
    "Gate results",
    "Outcome",
    "Fail reason",
    "Next action",
    "Next action date",
    "Flags",
]

GATE_ORDER = ["gate1", "gate2", "gate3", "gate4", "gate5", "filter3"]

PHASE2_STATUSES = {
    "PHASE2_SCOPED",
    "AWAITING_PHASE2_APPROVAL",
    "ESCALATED_TO_VICTORIA",
    "APPROVED",
    "CAPTURE_BRIEF_DRAFTED",
    "DOCS_DOWNLOADED",
    "CALENDARED",
    "ACTIVE",
}


def _format_date(value: str | None) -> str:
    if not value:
        return "UNVERIFIED"
    try:
        return datetime.fromisoformat(value).date().isoformat()
    except ValueError:
        return value


def _cpv_display(row: dict) -> str:
    primary = row["cpv_primary"] or "UNVERIFIED"
    additional = json.loads(row["cpv_additional"]) if row["cpv_additional"] else []
    additional = [c for c in additional if c != row["cpv_primary"]]
    text = f"{primary} (primary)"
    if row["cpv_primary_inferred"]:
        text += " [inferred]"
    if additional:
        text += f"; additional: {', '.join(additional)}"
    return text


def _next_action(row: dict) -> tuple[str, str]:
    """Returns (next action, next action date)."""
    status = row["status"]
    outcome = row["headline_outcome"]
    owner = row["owner"] or "unassigned"

    if outcome == "MONITOR":
        return (
            "Watch for re-tender or next stage",
            row["future_notice_date"] or row["deadline"] or "UNVERIFIED",
        )
    if outcome == "FAIL":
        return "None, out of scope", ""
    if status == "AWAITING_PHASE2_APPROVAL" and outcome in ("FLAG", "MAYBE"):
        return "Review Phase 2 read, consider Victoria decision", ""
    if status == "TO_REVIEW":
        return f"Awaiting {owner}'s Phase 1 approval", ""
    if status in PHASE2_STATUSES:
        return f"In Phase 2, status: {status}", ""
    return "Awaiting Phase 1 triage", ""


def _fetch_tracker_rows(conn: sqlite3.Connection) -> list[dict]:
    notices = conn.execute(
        """
        SELECT n.*, tr.id AS triage_run_id, tr.headline_gate, tr.headline_outcome, tr.headline_reason
        FROM notices n
        LEFT JOIN (
            SELECT notice_id, MAX(id) AS max_id FROM triage_runs GROUP BY notice_id
        ) latest ON latest.notice_id = n.id
        LEFT JOIN triage_runs tr ON tr.id = latest.max_id
        """
    ).fetchall()

    run_ids = [n["triage_run_id"] for n in notices if n["triage_run_id"] is not None]
    gate_results_by_run: dict[int, dict[str, tuple[str, str]]] = {}
    if run_ids:
        placeholders = ",".join("?" for _ in run_ids)
        for gr in conn.execute(
            f"SELECT * FROM gate_results WHERE triage_run_id IN ({placeholders})", run_ids
        ).fetchall():
            gate_results_by_run.setdefault(gr["triage_run_id"], {})[gr["gate_number"]] = (
                gr["outcome"],
                gr["reason"],
            )

    rows = []
    for n in notices:
        row = dict(n)
        row["gate_results"] = gate_results_by_run.get(n["triage_run_id"], {})
        rows.append(row)
    return rows


def _build_row_values(row: dict) -> list:
    gate_results = row["gate_results"]
    gate_summary = "; ".join(
        f"{g}:{gate_results[g][0]}" for g in GATE_ORDER if g in gate_results
    )
    flags = ", ".join(g for g in GATE_ORDER if g in gate_results and gate_results[g][0] in ("FLAG", "MAYBE"))
    outcome = row["headline_outcome"] or "PENDING"
    fail_reason = row["headline_reason"] if outcome != "PASS" else ""
    next_action, next_action_date = _next_action(row)

    return [
        row["ref"],
        _format_date(row["first_seen_at"]),
        row["title"],
        row["buyer"] or "UNVERIFIED",
        row["owner"] or "",
        row["source"],
        row["indicative_value"] or "UNVERIFIED",
        _cpv_display(row),
        row["uk_stage"],
        gate_summary,
        outcome,
        fail_reason,
        next_action,
        next_action_date,
        flags,
    ]


def _write_sheet(wb: Workbook, title: str, rows: list[dict]) -> Worksheet:
    ws = wb.create_sheet(title=title[:31])  # Excel sheet name limit
    ws.append(COLUMNS)
    for row in rows:
        ws.append(_build_row_values(row))
    return ws


def _write_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Legend and method")
    ws.append(["Savvy Scout Phase A tracker: legend and method"])
    ws.append([])
    ws.append(["Gate order (always all six run, no short-circuit):"])
    for key, name in [
        ("gate1", "Buyer sector and owner"),
        ("gate2", "Type of work (with CPV evidence)"),
        ("gate3", "Notice stage"),
        ("gate4", "Window"),
        ("gate5", "Framework status"),
        ("filter3", "Scale and incumbents (separately agreed rule, 15 June 2026, toggleable)"),
    ]:
        ws.append([key, name])
    ws.append([])
    ws.append(["Outcome values: PASS, FAIL, FLAG, MONITOR, INFERRED_FIT, MAYBE (Gate 3 PME route-not-yet-decided)."])
    ws.append(["Outcome is the headline: the first non-PASS gate result in gate order."])
    ws.append([])
    ws.append(["CPV lists sourced from the Kanvesh scouting skill Section 4, verified against Home Office SCBP notice 039639-2026."])
    ws.append(["Gate 5 is judged on the primary CPV code (tender.classification, or the first item classification, "
                "or, when neither is present, the first additionalClassifications entry, marked [inferred])."])
    ws.append(["Additional CPV codes are recorded for reference and noted only when they conflict with the primary outcome."])
    ws.append([])
    ws.append(["All figures are as stated in the source notice. Anything not present in the source is shown as UNVERIFIED, never invented."])
    ws.append(["All outcomes are rule-based Phase 1 triage results, not a bid/no-bid decision. Bid/no-bid belongs to Victoria Milan, Bid Director."])
    ws.append([])
    ws.append(["Sheet population rules:"])
    ws.append(["Phase 1 - Flags: headline outcome FLAG or MAYBE."])
    ws.append(["To review: headline outcome PASS or INFERRED_FIT, queued for or awaiting the owner's Phase 2 approval click "
                "(PASS/FLAG/MAYBE all skip the Phase 1 owner queue and go straight to the automated Phase 2 scope read)."])
    ws.append(["Handoffs - <owner>: that owner's rows from Flags and To review combined."])
    ws.append(["Closed or awarded: headline outcome FAIL from Gate 4 (window)."])
    ws.append(["Out of scope - no owner: headline outcome FAIL from any other gate."])
    ws.append(["Phase 2 - Pipeline: any notice from PHASE2_SCOPED onward, regardless of headline outcome."])


def export_tracker(conn: sqlite3.Connection, output_path: str) -> str:
    rows = _fetch_tracker_rows(conn)

    flags_rows = [r for r in rows if r["headline_outcome"] in ("FLAG", "MAYBE")]
    to_review_rows = [
        r
        for r in rows
        if r["headline_outcome"] in ("PASS", "INFERRED_FIT")
        and r["status"] in ("PHASE2_SCOPED", "AWAITING_PHASE2_APPROVAL")
    ]
    closed_or_awarded_rows = [
        r for r in rows if r["headline_outcome"] == "FAIL" and r["headline_gate"] == "gate4"
    ]
    out_of_scope_rows = [
        r
        for r in rows
        if r["headline_outcome"] == "FAIL" and r["headline_gate"] != "gate4"
    ]
    phase2_rows = [r for r in rows if r["status"] in PHASE2_STATUSES]

    owners = sorted({r["owner"] for r in rows if r["owner"]})

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    _write_sheet(wb, "Phase 1 - Flags", flags_rows)
    _write_sheet(wb, "To review", to_review_rows)
    for owner in owners:
        owner_rows = [r for r in (flags_rows + to_review_rows) if r["owner"] == owner]
        _write_sheet(wb, f"Handoffs - {owner}", owner_rows)
    _write_sheet(wb, "Closed or awarded", closed_or_awarded_rows)
    _write_sheet(wb, "Out of scope - no owner", out_of_scope_rows)
    _write_sheet(wb, "Phase 2 - Pipeline", phase2_rows)
    _write_legend_sheet(wb)

    wb.save(output_path)
    return output_path
