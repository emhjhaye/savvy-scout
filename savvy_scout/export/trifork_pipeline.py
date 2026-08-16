import os
import re
import sqlite3
from copy import copy
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from savvy_scout.escalation.context import MISSING, OWNER_NAMES, build_context

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "artifacts" / "pipeline_tracker_template.xlsx"
TRACKER_SHEETS = ("Pass", "Flag", "Fail")
CLEAR_SHEETS = TRACKER_SHEETS + ("Pass to Kanvesh", "Pass to NHS")
HEADERS = (
    "REF #", "DATE SPOTTED", "OPPORTUNITY TITLE", "BUYER", "BUYER TYPE", "SECTOR",
    "SOURCE", "NOTICE TYPE", "INDICATIVE VALUE", "CPV CODES",
    "SUBMISSION/ENGAGEMENT DEADLINE", "TRIAGE STATUS", "CAPABILITY FIT",
    "FRAMEWORK STATUS", "FILTER FLAGS (1/2/3)", "REASON / NOTES", "NEXT ACTION",
    "NEXT ACTION DATE", "OPEN FLAGS FOR VICTORIA",
    # 2026-08-16, explicit request: quick-reference links to the actual
    # generated documents, computed relative to the tracker's own folder so
    # they still resolve once both are pulled onto a local machine together.
    "INTERNAL ADDENDUM", "CAPTURE BRIEF",
)
TOTAL_COLUMNS = len(HEADERS)


def _owner_reviewed_notice_ids(conn: sqlite3.Connection, owner: str | None = None) -> list[int]:
    # owner, if given (e.g. "Mark"), restricts to that owner's own decisions
    # only -- explicit request (2026-08-16): Mark's tracker export must not
    # include Kanvesh's or Hammad's escalations/rejections.
    names = (owner,) if owner else OWNER_NAMES
    placeholders = ",".join("?" for _ in names)
    rows = conn.execute(
        f"SELECT notice_id, MAX(id) latest_id FROM status_history "
        f"WHERE from_status = 'AWAITING_PHASE2_APPROVAL' "
        f"AND to_status IN ('ESCALATED_TO_VICTORIA', 'REJECTED') "
        f"AND changed_by IN ({placeholders}) "
        f"AND EXISTS (SELECT 1 FROM phase2_assessments p WHERE p.notice_id = status_history.notice_id) "
        f"GROUP BY notice_id ORDER BY latest_id",
        names,
    ).fetchall()
    return [row["notice_id"] for row in rows]


def _decision_target(conn, notice_id):
    placeholders = ",".join("?" for _ in OWNER_NAMES)
    return conn.execute(
        f"SELECT to_status, reason FROM status_history WHERE notice_id = ? "
        f"AND from_status = 'AWAITING_PHASE2_APPROVAL' "
        f"AND to_status IN ('ESCALATED_TO_VICTORIA', 'REJECTED') "
        f"AND changed_by IN ({placeholders}) ORDER BY id DESC LIMIT 1",
        (notice_id, *OWNER_NAMES),
    ).fetchone()


def _target_sheet(context, decision):
    if decision["to_status"] == "REJECTED":
        return "Fail"
    return "Pass" if context["ai_read"]["overall"] == "PURSUE" else "Flag"


def _format_date(value):
    if value in (None, "", MISSING):
        return MISSING
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return parsed.strftime("%d/%m/%Y")
    except ValueError:
        try:
            return datetime.strptime(str(value)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except ValueError:
            return str(value)


def _buyer_type(value):
    if value in (None, "", MISSING):
        return "Unconfirmed — verify from the notice"
    mappings = {
        "public authority - central government": "Central government public authority",
        "public authority - sub-central government": "Sub-central public authority",
        "private sector": "Private-sector buyer",
    }
    return mappings.get(str(value).casefold(), str(value))


def _notice_type(context):
    stage = context["uk_stage"]
    label = {
        "UK1": "UK1 Pipeline notice",
        "UK2": "UK2 Preliminary Market Engagement",
        "UK3": "UK3 Planned procurement notice",
        "UK4": "UK4 Tender notice",
        "UK5": "UK5 Award notice",
    }.get(stage)
    if label:
        return label
    return context["notice_type"] if context["notice_type"] != MISSING else "Unconfirmed — verify notice type"


def _format_value(context):
    if context["value_estimate"] == MISSING:
        return "Not stated"
    try:
        amount = Decimal(str(context["value_estimate"]))
        if amount == 0:
            return "Not stated"
        formatted = f"{amount:,.0f}" if amount == amount.to_integral() else f"{amount:,.2f}"
    except (InvalidOperation, ValueError):
        return str(context["value_estimate"])
    symbol = {"GBP": "£", "EUR": "€", "USD": "$"}.get(context["currency"])
    return f"{symbol}{formatted}" if symbol else f"{formatted} {context['currency']}"


def _assessment_text(context):
    reasoning = context["ai_read"].get("per_field_reasoning", {})
    rating = context["ai_read"]["capability_fit"]
    return f"{rating}: {reasoning.get('capability_fit', MISSING)}"


def _first_sentences(value, count=1, limit=240):
    text = " ".join(str(value or MISSING).split())
    sentences = re.split(r"(?<=[.!?])\s+", text)
    result = " ".join(sentences[:count])
    if len(result) <= limit:
        return result
    shortened = result[: limit - 1].rsplit(" ", 1)[0].rstrip(" ,;:-")
    return shortened + "…"


def _source_url(context):
    if context["source_portal"].casefold() == "find a tender" and re.fullmatch(
        r"\d{6}-\d{4}", context["notice_reference"]
    ):
        return f"https://www.find-tender.service.gov.uk/Notice/{context['notice_reference']}"
    return context["notice_url"]


def _filter_flags(context):
    reasoning = context["ai_read"].get("per_field_reasoning", {})
    combined = " ".join([
        reasoning.get("capability_fit", ""), reasoning.get("right_to_win", ""),
        *(_item_text(value) for value in context["blockers_risks"]),
    ]).casefold()
    flags = []
    if context["ai_read"]["capability_fit"] == "LOW" and any(
        phrase in combined for phrase in ("capability gap", "not aligned", "lacks", "no plausible")
    ):
        flags.append("Filter 1 (capability/market category mismatch)")
    if any(phrase in combined for phrase in ("security clearance", "sc clearance", "dv clearance")):
        flags.append("Filter 2 (UK security clearance unconfirmed)")
    scale_gate = next(
        (gate for gate in context["gate_outcomes"] if "scale" in gate["gate_name"].casefold()), None
    )
    if scale_gate and scale_gate["result"] in ("FAIL", "FLAG"):
        flags.append(f"Filter 3 ({scale_gate['reason']})")
    return "; ".join(flags) if flags else "Clean"


def _item_text(value):
    if not isinstance(value, dict):
        return str(value or "")
    return " ".join(str(item) for item in value.values() if item)


def _flag_reason(context, decision):
    if decision["to_status"] == "REJECTED":
        return (
            f"FAIL after owner Phase 2 review. {context['ai_read'].get('per_field_reasoning', {}).get('overall', MISSING)} "
            f"Owner reason: {decision['reason'] or MISSING}."
        )
    reasoning = context["ai_read"].get("per_field_reasoning", {})
    parts = [
        f"FLAG — Trifork capability fit {context['ai_read']['capability_fit']}: "
        f"{_first_sentences(reasoning.get('capability_fit'), 1, 180)}",
        f"Right to win {context['ai_read']['right_to_win']}: "
        f"{_first_sentences(reasoning.get('right_to_win'), 1, 150)}",
    ]
    if context["framework_status"] != MISSING:
        framework = context["framework_status"]
        if "route not yet decided" in framework.casefold():
            framework = "Route not yet decided."
        elif "unclear" in framework.casefold():
            framework = "Framework status unconfirmed."
        else:
            framework = _first_sentences(framework, 1, 120)
        parts.append(framework)
    return " ".join(parts)


def _deadline_state(context):
    if context["submission_deadline"] == MISSING:
        return None, False
    try:
        deadline = datetime.strptime(context["submission_deadline"], "%Y-%m-%d").date()
        return deadline, deadline >= date.today()
    except ValueError:
        return None, False


def _next_action(context, decision):
    if decision["to_status"] == "REJECTED":
        return "Record the owner Phase 2 rejection and take no further capture action.", MISSING
    deadline, open_window = _deadline_state(context)
    formatted = deadline.strftime("%d/%m/%Y") if deadline else MISSING
    if open_window and context["uk_stage"] == "UK2":
        return f"Subject to Victoria's approval, respond to the market engagement by {formatted}.", formatted
    if open_window:
        return f"Victoria to decide GO / NO-GO; if GO, begin capture before the {formatted} submission deadline.", formatted
    if context["uk_stage"] == "UK2":
        return "Market engagement deadline has passed; monitor for the formal procurement notice and confirm continued interest.", formatted
    return "Confirm whether a replacement or amended procurement stage is open; otherwise close as expired.", formatted


def _victoria_question(context, decision):
    if decision["to_status"] == "REJECTED":
        return MISSING
    reasoning = context["ai_read"].get("per_field_reasoning", {})
    fit = _first_sentences(reasoning.get("capability_fit"), 1, 105)
    return (
        f"Does Victoria approve pursuing {context['title']} despite "
        f"{context['ai_read']['capability_fit']} capability fit and "
        f"{context['ai_read']['right_to_win']} right to win? Key issue: {fit}"
    )


def _row_values(context, decision):
    next_action, action_date = _next_action(context, decision)
    return [
        context["notice_reference"],
        _format_date(context["published_date"]),
        context["title"],
        context["buyer"],
        _buyer_type(context["buyer_org_type"]),
        context["sector"],
        context["source_portal"],
        _notice_type(context),
        _format_value(context),
        "; ".join(context["cpv_codes"]) or MISSING,
        _format_date(context["submission_deadline"]),
        _target_sheet(context, decision).upper(),
        _assessment_text(context),
        context["framework_status"],
        _filter_flags(context),
        _flag_reason(context, decision),
        next_action,
        action_date,
        _victoria_question(context, decision),
        "Not yet generated",  # INTERNAL ADDENDUM -- overwritten with a real link below if one exists
        "Not yet generated",  # CAPTURE BRIEF -- overwritten with a real link below if one exists
    ]


def _brief_link(conn: sqlite3.Connection, notice_id: int, brief_type: str, tracker_dir: Path) -> str | None:
    """Returns a path to the recorded brief file, relative to the tracker's
    own folder, so the hyperlink still resolves once both are copied onto a
    local machine together (see export_victoria_package, which puts the
    tracker and the Addendum/Brief folder under the same root)."""
    row = conn.execute(
        "SELECT docx_path FROM escalation_briefs WHERE notice_id = ? AND brief_type = ? "
        "ORDER BY id DESC LIMIT 1",
        (notice_id, brief_type),
    ).fetchone()
    if not row or not row["docx_path"]:
        return None
    try:
        return os.path.relpath(row["docx_path"], start=str(tracker_dir))
    except ValueError:
        return row["docx_path"]  # different drive on Windows -- fall back to the raw path


def _snapshot_styles(workbook):
    snapshots = {}
    for sheet_name in TRACKER_SHEETS:
        sheet = workbook[sheet_name]
        source_row = 3
        snapshots[sheet_name] = [copy(sheet.cell(source_row, column)._style) for column in range(1, TOTAL_COLUMNS + 1)]
    return snapshots


def _clear_sample_rows(workbook):
    for sheet_name in CLEAR_SHEETS:
        sheet = workbook[sheet_name]
        for row in range(3, sheet.max_row + 1):
            for column in range(1, TOTAL_COLUMNS + 1):
                sheet.cell(row, column).value = None
                sheet.cell(row, column).hyperlink = None


def _template_or_existing(output: Path):
    if output.exists():
        workbook = load_workbook(output)
        headers = tuple(workbook["Flag"].cell(2, column).value for column in range(1, TOTAL_COLUMNS + 1))
        if headers == HEADERS:
            return workbook, False
    return load_workbook(TEMPLATE_PATH), True


def update_trifork_pipeline(
    conn: sqlite3.Connection, output_path: str, owner: str | None = None
) -> dict[str, int | str]:
    output = Path(output_path)
    workbook, fresh = _template_or_existing(output)
    styles = _snapshot_styles(workbook)
    if fresh:
        _clear_sample_rows(workbook)

    existing = {}
    for sheet_name in TRACKER_SHEETS:
        sheet = workbook[sheet_name]
        for row in range(3, sheet.max_row + 1):
            reference = sheet.cell(row, 1).value
            if reference:
                existing[str(reference)] = (sheet_name, row)

    inserted = updated = skipped = 0
    for notice_id in _owner_reviewed_notice_ids(conn, owner):
        context = build_context(conn, notice_id)
        if any(context[key] == MISSING for key in ("notice_reference", "title", "buyer", "sector", "owner_name")):
            skipped += 1
            continue
        decision = _decision_target(conn, notice_id)
        target_name = _target_sheet(context, decision)
        reference = context["notice_reference"]
        old = existing.get(reference)
        if old:
            old_sheet, old_row = old
            if old_sheet != target_name:
                for column in range(1, TOTAL_COLUMNS + 1):
                    workbook[old_sheet].cell(old_row, column).value = None
                old = None
        if old:
            row_number = old[1]
            updated += 1
        else:
            target = workbook[target_name]
            occupied = [row for row in range(3, target.max_row + 1) if target.cell(row, 1).value]
            row_number = max(occupied, default=2) + 1
            inserted += 1
        sheet = workbook[target_name]
        for column, value in enumerate(_row_values(context, decision), start=1):
            cell = sheet.cell(row_number, column)
            cell._style = copy(styles[target_name][column - 1])
            cell.value = value
            base_font = copy(cell.font)
            cell.font = Font(
                name="Calibri", size=11, bold=base_font.bold, italic=base_font.italic,
                color="0563C1" if column == 7 else "000000",
                underline="single" if column == 7 else None,
            )
            if column in (13, 16, 17, 19):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            source_url = _source_url(context)
            if column == 7 and source_url != MISSING:
                cell.hyperlink = source_url
            if column == 20:  # INTERNAL ADDENDUM
                link = _brief_link(conn, notice_id, "INTERNAL_ADDENDUM", output.parent)
                if link:
                    cell.value = "Open Internal Addendum"
                    cell.hyperlink = link
                    cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
            if column == 21:  # CAPTURE BRIEF
                link = _brief_link(conn, notice_id, "CAPTURE_BRIEF", output.parent)
                if link:
                    cell.value = "Open Capture Brief"
                    cell.hyperlink = link
                    cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        sheet.row_dimensions[row_number].height = 105
        existing[reference] = (target_name, row_number)

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(output)
    return {
        "inserted": inserted, "updated": updated, "skipped": skipped,
        "total": len(existing), "output_path": str(output),
    }


def update_configured_trifork_pipeline(conn: sqlite3.Connection):
    output_path = os.environ.get("TRIFORK_PIPELINE_OUTPUT_PATH")
    owner = os.environ.get("TRIFORK_PIPELINE_OWNER") or None
    return update_trifork_pipeline(conn, output_path, owner) if output_path else None
